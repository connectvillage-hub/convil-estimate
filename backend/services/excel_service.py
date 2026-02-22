import io
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from models.estimate import EstimateRequest, EstimateResult

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'estimate_template.xlsx')

# 색상 상수
PRIMARY_COLOR = '2E75B6'
PRIMARY_DARK = '1a4f80'
PRIMARY_LIGHT = 'EBF3FB'
GRAY_BORDER = 'CCCCCC'
GRAY_LIGHT = 'F5F5F5'
WHITE = 'FFFFFF'

# 테두리 스타일
def thin_border():
    side = Side(style='thin', color=GRAY_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def bottom_border(color=GRAY_BORDER):
    return Border(bottom=Side(style='thin', color=color))

def num_fmt(ws, cell_ref, value):
    """숫자 포맷 적용 (#,##0)"""
    cell = ws[cell_ref]
    cell.value = value
    cell.number_format = '#,##0'
    return cell


def generate_excel(req: EstimateRequest, result: EstimateResult) -> bytes:
    # 템플릿 파일이 있으면 사용, 없으면 새로 생성
    if os.path.exists(TEMPLATE_PATH):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        _fill_template(ws, req, result)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "견적서"
        _create_from_scratch(ws, req, result)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _fill_template(ws, req: EstimateRequest, result: EstimateResult):
    """기존 템플릿에 데이터만 채우기"""
    # 견적일자
    ws['B7'] = req.estimateDate

    # 항목 채우기 (B10~F29, 최대 20행)
    items = result.itemDetails
    for i, item in enumerate(items[:20]):
        row = 10 + i
        ws[f'B{row}'] = item.scope
        ws[f'C{row}'] = item.item
        ws[f'D{row}'] = item.quantity
        ws[f'E{row}'].value = item.unitCost
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'F{row}'].value = item.cost
        ws[f'F{row}'].number_format = '#,##0'

    # 합계
    ws['F30'].value = result.subtotal
    ws['F30'].number_format = '#,##0'
    ws['F31'].value = result.discount
    ws['F31'].number_format = '#,##0'
    ws['F32'].value = result.total
    ws['F32'].number_format = '#,##0'
    ws['F33'].value = result.vat
    ws['F33'].number_format = '#,##0'
    ws['E35'].value = result.finalAmount
    ws['E35'].number_format = '#,##0'


def _create_from_scratch(ws, req: EstimateRequest, result: EstimateResult):
    """처음부터 견적서 Excel 생성"""

    # ── 열 너비 설정 ──
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 36
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    # ── 행 높이 설정 ──
    for r in range(1, 45):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[30].height = 22
    ws.row_dimensions[31].height = 22
    ws.row_dimensions[32].height = 22
    ws.row_dimensions[33].height = 22
    ws.row_dimensions[35].height = 30

    # ── 헤더 영역 (Row 1~3) ──
    ws.merge_cells('B2:D2')
    logo_cell = ws['B2']
    logo_cell.value = 'CONVIL DESIGN'
    logo_cell.font = Font(name='Arial', size=22, bold=True, color=PRIMARY_COLOR)
    logo_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 회사 정보 (우측)
    ws['E2'] = '컨빌디자인'
    ws['E2'].font = Font(name='Arial Unicode MS', size=14, bold=True)
    ws['E2'].alignment = Alignment(horizontal='right', vertical='center')

    ws['F2'] = '견 적 서'
    ws['F2'].font = Font(name='Arial Unicode MS', size=18, bold=True, color=PRIMARY_COLOR)
    ws['F2'].alignment = Alignment(horizontal='right', vertical='center')

    # Row 3: 구분선
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}3'].fill = PatternFill('solid', fgColor=PRIMARY_COLOR)
    ws.row_dimensions[3].height = 4

    # ── 회사 정보 Row 4~6 ──
    ws['E4'] = '컨빌디자인'
    ws['E4'].font = Font(name='Arial Unicode MS', size=11, bold=True)
    ws['E4'].alignment = Alignment(horizontal='right')

    ws['E5'] = 'www.convil.net'
    ws['E5'].font = Font(name='Malgun Gothic', size=10, color='555555')
    ws['E5'].alignment = Alignment(horizontal='right')

    ws['E6'] = '대표자 박진하 (인)'
    ws['E6'].font = Font(name='Arial Unicode MS', size=10, color='555555')
    ws['E6'].alignment = Alignment(horizontal='right')

    # 고객 정보 Row 4~6 (좌측)
    ws['B4'] = '견적일자'
    ws['B4'].font = Font(name='Malgun Gothic', size=9, color='888888')
    ws['C4'] = req.estimateDate
    ws['C4'].font = Font(name='Malgun Gothic', size=9, bold=True)

    ws['B5'] = '고객명'
    ws['B5'].font = Font(name='Malgun Gothic', size=9, color='888888')
    ws['C5'] = req.customerName
    ws['C5'].font = Font(name='Malgun Gothic', size=9, bold=True)

    ws['B6'] = '프로젝트'
    ws['B6'].font = Font(name='Malgun Gothic', size=9, color='888888')
    ws['C6'] = req.projectName
    ws['C6'].font = Font(name='Malgun Gothic', size=9, bold=True)

    # ── 견적일자 Row 7 (큰 글씨) ──
    ws['B7'] = req.estimateDate
    ws['B7'].font = Font(name='Malgun Gothic', size=10)
    ws.row_dimensions[7].height = 6

    # ── 구분선 Row 8 ──
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}8'].fill = PatternFill('solid', fgColor='EEEEEE')
    ws.row_dimensions[8].height = 4

    # ── 항목 헤더 Row 9 ──
    headers = [('B', 'Scope'), ('C', 'Item'), ('D', 'QTY'), ('E', 'Unit Cost'), ('F', 'Cost')]
    for col, label in headers:
        cell = ws[f'{col}9']
        cell.value = label
        cell.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=PRIMARY_COLOR)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()

    # ── 항목 행 Row 10~29 ──
    items = result.itemDetails
    for i in range(20):
        row = 10 + i
        bg = PatternFill('solid', fgColor=GRAY_LIGHT if i % 2 == 0 else WHITE)

        for col in ['B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.fill = bg
            cell.border = thin_border()
            cell.font = Font(name='Malgun Gothic', size=9)

        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'E{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'F{row}'].alignment = Alignment(horizontal='right', vertical='center')

        if i < len(items):
            item = items[i]
            ws[f'B{row}'].value = item.scope
            ws[f'C{row}'].value = item.item
            ws[f'D{row}'].value = item.quantity
            ws[f'E{row}'].value = item.unitCost
            ws[f'E{row}'].number_format = '#,##0'
            ws[f'F{row}'].value = item.cost
            ws[f'F{row}'].number_format = '#,##0'

    # ── 합계 영역 Row 30~33 ──
    summary_items = [
        (30, 'Subtotal', result.subtotal),
        (31, 'Discount', result.discount),
        (32, 'Total', result.total),
        (33, 'VAT (10%)', result.vat),
    ]

    for row, label, value in summary_items:
        ws.row_dimensions[row].height = 22
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'B{row}'].fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
        ws[f'B{row}'].border = thin_border()
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')

        ws[f'F{row}'].value = value
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].font = Font(name='Arial', size=10, bold=(label in ('Subtotal', 'Total')))
        ws[f'F{row}'].fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
        ws[f'F{row}'].border = thin_border()
        ws[f'F{row}'].alignment = Alignment(horizontal='right', vertical='center')

        # 가운데 셀 채우기
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
            ws[f'{col}{row}'].border = thin_border()

    # ── 최종 합계 Row 35 ──
    ws.merge_cells('B35:D35')
    ws['B35'] = '최종 합계금액 (VAT 포함)'
    ws['B35'].font = Font(name='Malgun Gothic', size=11, bold=True, color=WHITE)
    ws['B35'].fill = PatternFill('solid', fgColor=PRIMARY_COLOR)
    ws['B35'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B35'].border = thin_border()

    ws.merge_cells('E35:F35')
    ws['E35'].value = result.finalAmount
    ws['E35'].number_format = '#,##0'
    ws['E35'].font = Font(name='Arial', size=14, bold=True, color=PRIMARY_COLOR)
    ws['E35'].fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
    ws['E35'].alignment = Alignment(horizontal='right', vertical='center')
    ws['E35'].border = thin_border()

    # ── 안내 문구 Row 37~38 ──
    ws.row_dimensions[37].height = 16
    ws.row_dimensions[38].height = 16
    ws['B37'] = '※ 본 견적서는 발행일로부터 30일간 유효합니다.'
    ws['B37'].font = Font(name='Malgun Gothic', size=9, color='555555')
    ws['B38'] = '※ 계약금 입금 후 작업이 시작되며, 작업 완료 후 잔금을 납부하여 주시기 바랍니다.'
    ws['B38'].font = Font(name='Malgun Gothic', size=9, color='555555')

    # ── 연락처 Row 41 ──
    ws.row_dimensions[41].height = 16
    ws['B41'] = '컨빌디자인  |  대표 박진하  |  www.convil.net'
    ws['B41'].font = Font(name='Malgun Gothic', size=9, color='888888')

    # 인쇄 영역 설정
    ws.print_area = 'A1:G42'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
