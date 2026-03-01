import io
from openpyxl import load_workbook
from models.schemas.material import MaterialCreate
from typing import List


# 엑셀 헤더 → 필드명 매핑 (한글/영문 모두 지원)
COLUMN_MAP = {
    "자재명": "name",
    "재료명": "name",
    "name": "name",
    "카테고리": "category",
    "분류": "category",
    "category": "category",
    "단위": "unit",
    "unit": "unit",
    "단가": "unit_price",
    "가격": "unit_price",
    "unit_price": "unit_price",
    "규격": "spec",
    "spec": "spec",
    "제조사": "manufacturer",
    "manufacturer": "manufacturer",
    "메모": "memo",
    "비고": "memo",
    "memo": "memo",
    "할증률": "waste_percent",
    "할증률(%)": "waste_percent",
    "waste_percent": "waste_percent",
}


def parse_material_excel(contents: bytes) -> List[MaterialCreate]:
    wb = load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active

    # 첫 행에서 헤더 읽기
    headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip()
        field = COLUMN_MAP.get(val, None)
        headers.append(field)

    if "name" not in headers:
        raise ValueError("'자재명' 또는 'name' 열이 필요합니다.")

    materials = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                data[headers[i]] = val

        # 빈 행 건너뛰기
        if not data.get("name"):
            continue

        # 타입 변환
        data["name"] = str(data["name"]).strip()
        data["category"] = str(data.get("category", "기타") or "기타").strip()
        data["unit"] = str(data.get("unit", "m2") or "m2").strip()
        data["unit_price"] = float(data.get("unit_price", 0) or 0)
        data["spec"] = str(data.get("spec", "") or "").strip()
        data["manufacturer"] = str(data.get("manufacturer", "") or "").strip()
        data["memo"] = str(data.get("memo", "") or "").strip()
        data["waste_percent"] = float(data.get("waste_percent", 0) or 0)

        materials.append(MaterialCreate(**data))

    wb.close()
    return materials
